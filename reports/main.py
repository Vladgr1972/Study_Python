import os
from openpyxl import load_workbook


file_name_reports = 'detail.xlsx'
path_name = "d:\\temp\\2022\\reports\\in"


def start_reports(file_name_reports):
    date_reports, list_atm = load_date(file_name_reports)
    dict_atm = create_list_atm()
    not_used_atm = check_dict_atm(list_atm, dict_atm)
    if len(not_used_atm) != 0:
        print(not_used_atm)
    # print(dict_atm)
    output_date(list_atm, date_reports)


def check_dict_atm(list_atm, dict_atm):
    """ Checking that all ATMs are present in the report for the period """
    list_not_used_atm = []
    for el_list in list_atm:
        flag = False
        for el_dict in dict_atm:
            if not el_list[0] in dict_atm[el_dict] and not flag:
                flag = False
            else:
                flag += True
        if not flag:
            list_not_used_atm.append((el_list[0], el_list[1]))
    return list_not_used_atm


def load_date(file_name):
    """ Loading data from the file detail.xlsx and creating a list of data to be transmitted for further processing"""
    list_atm = []
    wb = load_workbook(file_name)
    list_sheets = wb.sheetnames
    sheet0 = wb[list_sheets[0]]
    sheet1 = wb[list_sheets[1]]
    for i in range(2, sheet0.max_row + 1):
        id_atm = sheet0.cell(row=i, column=2).value
        sum_proc = sheet0.cell(row=i, column=sheet0.max_column - 1).value + \
            sheet0.cell(row=i, column=sheet0.max_column).value
        sum_card = sheet1.cell(row=i, column=sheet1.max_column - 1).value + \
            sheet1.cell(row=i, column=sheet1.max_column).value
        sum_un_used_card = sum_card * (100 / sum_proc) - sum_card
        model_atm = sheet0.cell(row=i, column=6).value
        name_atm = sheet0.cell(row=i, column=5).value
        # print(f'{id_atm} {sum_proc} {sum_card} {sum_un_used_card}')
        list_atm.append((id_atm, name_atm, model_atm, sum_proc, sum_card, sum_un_used_card))
    date_reports = sheet0.cell(row=2, column=10).value.replace('.', '') + '_' + \
        sheet0.cell(row=2, column=11).value.replace('.', '')
    return date_reports, list_atm


def check_file(file_name, path):
    os.chdir(path)
    if not os.path.isfile(file_name):
        print(f"The file {file_name} is missing in the directory {path}")
        flag = False
    else:
        flag = True
    return flag


def output_date(date_atm, date_reports):
    date_atm.sort(key=lambda x: x[3])
    # print(date_atm)
    # print(date_reports)


def create_list_atm(file_name='list_atm.xlsx', path='d:\\temp\\2022\\reports\\in'):    
    """ Create dict of atm from file list_atm """
    atm_dict = {}
    if check_file(file_name, path):
        wb = load_workbook(file_name)
        list_sheets = wb.sheetnames
        for elem in list_sheets:
            sheet = wb[elem]
            for i in range(2, sheet.max_row+1):
                if elem not in atm_dict:
                    atm_dict[elem] = (sheet.cell(row=i, column=1).value, )
                else:
                    atm_dict[elem] += (sheet.cell(row=i, column=1).value, )
    else:
        pass
    return atm_dict


if __name__ == '__main__':
    if check_file(file_name_reports, path_name):
        start_reports(file_name_reports)
    else:
        pass
